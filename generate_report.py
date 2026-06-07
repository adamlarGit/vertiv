from __future__ import annotations

import argparse
import shutil
import re
from datetime import date, datetime, time
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

from docxtpl import DocxTemplate
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "EXCEL" / "data.xlsx"
DEFAULT_TEMPLATE = ROOT / "TEMPLATE" / "TEMPLATE.docx"
DEFAULT_OUTPUT = ROOT / "OUTPUT" / "thermal_report.docx"

# Temporary report-level defaults. If these headers are later added to Excel,
# non-blank Excel values will override these constants row by row.
DEFAULT_CONSTANTS: dict[str, Any] = {
    "ptu.no": "PTU 09",
    "model": "",
    "rating": "415V",
}

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "v": "urn:schemas-microsoft-com:vml",
    "o": "urn:schemas-microsoft-com:office:office",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

for prefix, uri in NS.items():
    if prefix != "rel":
        ET.register_namespace(prefix, uri)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Render one Word report page per Excel row, preserving the source DOCX page objects."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Input Excel data file.")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="One-page DOCX template.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Final combined DOCX output.")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Defaults to the active sheet.")
    parser.add_argument(
        "--keep-pages",
        action="store_true",
        help="Keep intermediate rendered page DOCX files next to the final output for debugging.",
    )
    parser.add_argument(
        "--merge-engine",
        choices=("docxcompose", "word"),
        default="word",
        help="Merge rendered pages with docxcompose or Microsoft Word COM automation.",
    )
    return parser.parse_args()


def format_value(header: str, value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, datetime):
        if header == "date":
            return value.strftime("%d-%m-%Y")
        if header == "time":
            return value.strftime("%H:%M")
        return value.isoformat(sep=" ")

    if isinstance(value, date) and header == "date":
        return value.strftime("%d-%m-%Y")

    if isinstance(value, time):
        return value.strftime("%H:%M")

    if header == "humidity" and isinstance(value, (int, float)):
        return f"{value:.0%}" if 0 <= value <= 1 else str(value)

    return value


def set_nested(context: dict[str, Any], dotted_key: str, value: Any) -> None:
    parts = dotted_key.split(".")
    target = context
    for part in parts[:-1]:
        current = target.get(part)
        if not isinstance(current, dict):
            current = {}
            target[part] = current
        target = current
    target[parts[-1]] = value


def build_context(row: dict[str, Any]) -> dict[str, Any]:
    context: dict[str, Any] = {}

    for key, value in DEFAULT_CONSTANTS.items():
        set_nested(context, key, value)

    for key, value in row.items():
        if key is None:
            continue
        normalized_key = str(key).strip()
        if not normalized_key:
            continue

        formatted = format_value(normalized_key, value)
        if formatted == "" and normalized_key in DEFAULT_CONSTANTS:
            continue
        set_nested(context, normalized_key, formatted)

    return context


def load_rows(excel_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_idx, col).value for col in range(1, worksheet.max_column + 1)]
        if not any(value is not None for value in values):
            continue
        rows.append(dict(zip(headers, values)))

    if not rows:
        raise ValueError(f"No data rows found in {excel_path}")

    return rows


def render_page(template_path: Path, context: dict[str, Any], output_path: Path) -> None:
    template = DocxTemplate(str(template_path))
    template.render(context)
    template.save(output_path)


def combine_pages(page_paths: list[Path], output_path: Path) -> None:
    if not page_paths:
        raise ValueError("No rendered pages to combine.")

    from docx import Document
    from docxcompose.composer import Composer

    master = Document(str(page_paths[0]))
    composer = Composer(master)

    for page_path in page_paths[1:]:
        master.add_page_break()
        composer.append(Document(str(page_path)))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    composer.save(str(output_path))


def combine_pages_with_word(page_paths: list[Path], output_path: Path) -> None:
    if not page_paths:
        raise ValueError("No rendered pages to combine.")

    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Microsoft Word merge requires pywin32/win32com.") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    master_path = output_path.parent / f"{output_path.stem}.word_merge_working.docx"
    if master_path.exists():
        master_path.unlink()
    shutil.copy2(page_paths[0], master_path)

    word = win32com.client.DispatchEx("Word.Application")
    word.Visible = False
    word.DisplayAlerts = 0

    document = None
    try:
        document = word.Documents.Open(str(master_path.resolve()), ReadOnly=False, AddToRecentFiles=False)
        for page_path in page_paths[1:]:
            insert_range = document.Range(document.Content.End - 1, document.Content.End - 1)
            insert_range.InsertBreak(7)  # wdPageBreak
            insert_range = document.Range(document.Content.End - 1, document.Content.End - 1)
            insert_range.InsertFile(str(page_path.resolve()))

        document.SaveAs2(str(output_path.resolve()), FileFormat=12)  # wdFormatXMLDocument
    finally:
        if document is not None:
            try:
                document.Close(SaveChanges=False)
            except Exception:
                # Word can disconnect the COM object after SaveAs2 while the file is still saved.
                pass
        try:
            word.Quit()
        except Exception:
            pass
        if master_path.exists():
            master_path.unlink()

    if not output_path.exists() or output_path.stat().st_size == 0:
        raise RuntimeError(f"Word merge did not create a valid output file: {output_path}")


def make_flir_objects_independent(docx_path: Path) -> None:
    """Give every copied FLIR placeholder its own field names, shape ID, and WMF part.

    Word's manual copy/paste behavior creates unique VML object identities per page.
    docxcompose keeps repeated source relationships compact, which is visually valid
    but risky for FLIR Tools+ because multiple pages can point to the same field/media
    slot. This post-process makes each page-level placeholder independent again.
    """

    with ZipFile(docx_path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    document_xml = entries["word/document.xml"]
    rels_xml = entries["word/_rels/document.xml.rels"]
    document_root = ET.fromstring(document_xml)
    rels_root = ET.fromstring(rels_xml)

    # Keep FLIR DOCVARIABLE field names unique per report page/table.
    tables = document_root.findall(".//w:tbl", NS)
    for page_idx, table in enumerate(tables, start=1):
        field_idx = 0
        for instr_text in table.findall(".//w:instrText", NS):
            text = instr_text.text or ""
            if "DOCVARIABLE" not in text:
                continue
            field_idx += 1
            field_name = f"_Fd{800000000 + (page_idx * 100) + field_idx}"
            instr_text.text = re.sub(r"DOCVARIABLE\s+\S+", f"DOCVARIABLE {field_name}", text, count=1)

    # Keep VML shape IDs unique per FLIR placeholder object.
    shapes = document_root.findall(".//v:shape", NS)
    for shape_idx, shape in enumerate(shapes, start=1):
        shape.set("id", f"_x0000_i{3000 + shape_idx}")

    # Duplicate the VML preview WMF/media relationship per object, matching Word paste behavior.
    relationships = list(rels_root)
    rel_by_id = {rel.get("Id"): rel for rel in relationships}
    used_rids = {rel.get("Id") for rel in relationships if rel.get("Id")}
    next_rid = 1000

    def next_relationship_id() -> str:
        nonlocal next_rid
        while f"rId{next_rid}" in used_rids:
            next_rid += 1
        rid = f"rId{next_rid}"
        used_rids.add(rid)
        next_rid += 1
        return rid

    image_nodes = document_root.findall(".//v:imagedata", NS)
    for image_idx, image_node in enumerate(image_nodes, start=1):
        old_rid = image_node.get(f"{{{NS['r']}}}id")
        old_rel = rel_by_id.get(old_rid)
        if old_rel is None:
            continue

        old_target = old_rel.get("Target")
        old_type = old_rel.get("Type")
        if not old_target or not old_type:
            continue

        old_media_path = f"word/{old_target}"
        if old_media_path not in entries:
            continue

        extension = Path(old_target).suffix or ".wmf"
        new_target = f"media/flir_slot_{image_idx:03d}{extension}"
        new_media_path = f"word/{new_target}"
        entries[new_media_path] = entries[old_media_path]

        new_rid = next_relationship_id()
        new_rel = ET.Element(f"{{{NS['rel']}}}Relationship")
        new_rel.set("Id", new_rid)
        new_rel.set("Type", old_type)
        new_rel.set("Target", new_target)
        rels_root.append(new_rel)
        image_node.set(f"{{{NS['r']}}}id", new_rid)

    entries["word/document.xml"] = ET.tostring(document_root, encoding="utf-8", xml_declaration=True)
    entries["word/_rels/document.xml.rels"] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)

    temp_path = docx_path.with_suffix(".tmp.docx")
    with ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for name, content in entries.items():
            target.writestr(name, content)
    temp_path.replace(docx_path)


def generate_report(
    excel_path: Path,
    template_path: Path,
    output_path: Path,
    sheet_name: str | None,
    keep_pages: bool,
    merge_engine: str,
) -> None:
    rows = load_rows(excel_path, sheet_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if keep_pages:
        pages_dir = output_path.with_suffix("")
        if pages_dir.exists():
            shutil.rmtree(pages_dir)
        pages_dir.mkdir(parents=True)
        page_paths = []
        for idx, row in enumerate(rows, start=1):
            page_path = pages_dir / f"page_{idx:03d}.docx"
            render_page(template_path, build_context(row), page_path)
            page_paths.append(page_path)
        if merge_engine == "word":
            combine_pages_with_word(page_paths, output_path)
        else:
            combine_pages(page_paths, output_path)
            make_flir_objects_independent(output_path)
        return

    with TemporaryDirectory(dir=output_path.parent) as temp_dir:
        temp_path = Path(temp_dir)
        page_paths = []
        for idx, row in enumerate(rows, start=1):
            page_path = temp_path / f"page_{idx:03d}.docx"
            render_page(template_path, build_context(row), page_path)
            page_paths.append(page_path)
        if merge_engine == "word":
            combine_pages_with_word(page_paths, output_path)
        else:
            combine_pages(page_paths, output_path)
            make_flir_objects_independent(output_path)


def main() -> None:
    args = parse_args()
    generate_report(
        excel_path=args.excel,
        template_path=args.template,
        output_path=args.output,
        sheet_name=args.sheet,
        keep_pages=args.keep_pages,
        merge_engine=args.merge_engine,
    )
    print(f"Generated: {args.output}")


if __name__ == "__main__":
    main()
